import streamlit as st
import pandas as pd
import plotly.express as px
from datetime import datetime
import os
from openpyxl import load_workbook, Workbook
from PIL import Image
import gspread
from google.oauth2.service_account import Credentials

# --- [1. 기본 설정 및 데이터 구조] ---
EXCEL_FILE = "fire_inspection_log.xlsx"
SHEET_NAME = "부천성모병원_소방점검_데이터"

# 건물별 상세 층 구성 정보 (이용민 님 제공 데이터)
BUILDING_CONFIG = {
    "성모관(A동)": ["B1F", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "11F"],
    "성심관(L동)": ["B6F", "B6MF", "B5F", "B4F", "B3F", "B2F", "B1F", "1F", "2F", "3F", "4F", "5F", "6F", "7F", "8F", "9F", "10F", "PHF"],
    "성가정관(A동)": ["B1F", "1F", "2F", "3F", "4F", "5F", "6F"],
    "성요셉관(G동)": ["B2F", "B1F", "1F", "2F", "3F", "4F", "5F", "PHF"],
    "지하주차장(K동)": ["B4F", "B3F", "B2F", "B1F", "1F"],
    "주차타워(N동)": ["B4F", "B3F", "B2F", "B1F", "1F", "2F", "3F", "4F", "PHF"]
}

TOTAL_ITEMS = ["소화기구", "소화가스구역", "옥내소화전설비", "스프링클러설비", "자탐설비(감지기)", 
               "유도등설비", "비상조명등설비", "완강기", "구조대", "방열복", "공기호흡기", 
               "특피제연설비", "상가제연설비", "비상콘센트", "무선통신설비"]

# --- [2. 구글 시트 연결 함수] ---
def connect_google_sheet():
    try:
        # V75에서 제공된 보안 키 (ef13b193...)
        one_line_key = "-----BEGIN PRIVATE KEY-----\nMIIEvgIBADANBgkqhkiG9w0BAQEFAASCBKgwggSkAgEAAoIBAQDEu2iH7S0EChAv\nxlyqeFqMzoTGdLtVUOBGbuZYK+uibyqgFrKzzbfhGuebfZf2yq9Cb+LOFtFk74yg\nn8sUSSh/lpvQ9/ovJc2sDKK9EfGq75RXLr56dxuTLgJYSvnaQatyd/+q6HH3nrkK\nignp2jU4x/2xXRAVxKaTAjBZ9qRsK07sXXBy0KGAv0S2hDIwloDYWLZJykZJqomG\no/WBDu7VXanS80U8NXkUY04+D8WiRIRddam/yuZw7lDJIw0sovLPCnG9XofktmQi\nZAIeyp6v9SxwThgH/Ezfe3yYiAizUDyapRoo7fCufilyo7AaA8VzhR3f2nl3IFRa\n8WuLWETDAgMBAAECggEAE6Co+Kq6i4UfgUy2kRFp41N1XsFlsUjc31oQJCrZmv6h\nT5CEKrkzB8Ph6Jr6rtYg7y1vgOU45euaaaoXnkIARt51VD9xWAPbBbfLArcIF+hI\nMfyJZ3l7qE72T+3HCLX1j3mWaew/9pcA2iLUg17QYYZecrQYtztD7TXSMh+P3vS0\nIYGzJLdMag6pEsY9WHLfy9vtOsaIVQ73Lrz0a3IgilVJMdt5I9m1JVpLuxz68NUR\nuFTDSeMuG1kEEj5CAP0lOGdbAAA0Voz1zD5LDQaBvePDUo/+6p9ugvp8o1Py2BZu\nmqr4pDjEwk+1/L/Gcfdqg7mFEZ9wHcIcHT/BnICDqQKBgQD6f03RiJ39peDeG2LP\nuw915pA/3scYO1pCjyJMFt91LyeOsdBoqaw6vTFl5POBUuh3YpNfmnYjeTNoRe+g\nDxra5nCzqLO+S50RD96xE/A532PBYGK6ijA9dl2DR4SjjYiro5P9+NhRKf/HMr4J\nDyM0crkLSUrZaOHCPpr3NpQ4aQKBgQDJDcAYgcKeeiOZHZMyZEqt44/YIqDc/JXX\nlsY7SDf4zJt9ptEIVRyJQcfb9baGwm3S8KVsC9MHgOCTin1vTFo3weUnhkzEwLCx\n32dBXmwcF7VVH7mBSrp0rsAPwho75hY8czeVyeE8z3jpCMyHR7nYWnmgBY8HUDA0\n6cbtEWYOSwKBgAW13TpHmQJtQ6EMksYVh5ayeNtb4sIEqFwYEzoU02dCee5g01Vv\neMOmCHhALeTm1PHTx3RIeyT34FoiLgJCFeSjmDIgSGiwJJZ8/CChSbqoAwgngFis\Sck854jUmot38IOpKPiDMbs37+LEn1xpge3yJ5HAS/qEz5C0CKhWEC7xAoGBALFs\Kvr1d7O2gTebu8waaTSuvr+zP1YEqI7KBUFHZ+1tUkI+NfUudrJBlIdtfJfLeHNP\n3gtNG/thJc3Nxweyz2Ko/vo0MAbpLISra0lqc+kOmvz1tBR+x7A0jlTZq6D0WOcb\nng/XSacKR98N3r/68ewVW4N1pFSF1SS8sMmAMxmlAoGBAI73RTsF/u9WWqyhfUY9\nv8X8qNbt7Hlx2kRwyxQ4EZBWUcQpkIPBBxYXwnG1MCB4RBPFdaQdmyIE6f48weAe\nDjNIf2CD7ADF6edD1WI6+9Kp3dU3VpYcGvcot3/8Yh6Tcax/jDcG4EsapwnmjvXh\n2HEDRqLrm6ENvQk9Rmo2Xqr6\n-----END PRIVATE KEY-----\n"
        service_account_info = {
            "type": "service_account",
            "project_id": "round-booking-494300-s3",
            "private_key_id": "ef13b193ed296153d55ea7f29defc01969337073",
            "private_key": one_line_key,
            "client_email": "id-298@round-booking-494300-s3.iam.gserviceaccount.com",
            "client_id": "114249893845931311645",
            "auth_uri": "https://accounts.google.com/o/oauth2/auth",
            "token_uri": "https://oauth2.googleapis.com/token",
            "auth_provider_x509_cert_url": "https://www.googleapis.com/oauth2/v1/certs",
            "client_x509_cert_url": "https://www.googleapis.com/robot/v1/metadata/x509/id-298%40round-booking-494300-s3.iam.gserviceaccount.com"
        }
        credentials = Credentials.from_service_account_info(
            service_account_info,
            scopes=["https://www.googleapis.com/auth/spreadsheets", "https://www.googleapis.com/auth/drive"]
        )
        client = gspread.authorize(credentials)
        return client.open(SHEET_NAME).sheet1
    except:
        return None

# --- [3. 메인 UI 및 입력 폼] ---
st.set_page_config(page_title="부천성모병원 소방점검", layout="wide")

# 로고 상단 배치
try:
    logo_img = Image.open("logo.png")
    col_l, col_r = st.columns([1, 6])
    with col_l: st.image(logo_img, width=150)
    with col_r: st.markdown("<h1 style='margin-top: 15px;'>소방시설 실시간 점검 시스템</h1>", unsafe_allow_html=True)
except:
    st.title("🏥 소방시설 실시간 점검 시스템")

# 점검 기본 정보 입력
st.sidebar.header("📋 점검 기본 정보")
inspector = st.sidebar.text_input("점검자", value="이용민")
check_date = st.sidebar.date_input("점검 일자", datetime.now())
selected_bldg = st.sidebar.selectbox("건물 선택", list(BUILDING_CONFIG.keys()))
selected_floor = st.sidebar.selectbox("층수 선택", BUILDING_CONFIG[selected_bldg])
full_location = f"{selected_bldg} {selected_floor}"

# 점검 항목 라디오 버튼 배치
st.header(f"🔍 {full_location} 시설물 상태 체크")
results = {}
cols = st.columns(3)
for idx, item in enumerate(TOTAL_ITEMS):
    with cols[idx % 3]:
        results[item] = st.radio(f"**{item}**", ["양호", "불량"], key=f"check_{item}", horizontal=True)

# 비고란
issue_detail = st.text_area("📝 지적 내역 및 비고 (불량 사유 등)", height=100)

# 데이터 저장 버튼
if st.button("📊 점검 결과 저장 및 전송", use_container_width=True):
    new_row = [check_date.strftime("%Y-%m-%d"), inspector, full_location] + list(results.values()) + [issue_detail]
    
    # 1. 로컬 엑셀 백업
    try:
        if not os.path.exists(EXCEL_FILE):
            wb = Workbook(); ws = wb.active; ws.append(["일자", "점검자", "구역"] + TOTAL_ITEMS + ["비고"])
        else:
            wb = load_workbook(EXCEL_FILE); ws = wb.active
        ws.append(new_row); wb.save(EXCEL_FILE)
    except: pass

    # 2. 구글 시트 실시간 전송
    sheet = connect_google_sheet()
    if sheet:
        try:
            sheet.append_row(new_row)
            st.success("✅ 구글 스프레드시트에 성공적으로 저장되었습니다!")
            st.balloons()
        except Exception as e:
            st.error(f"시트 전송 중 오류 발생: {e}")
    else:
        st.warning("⚠️ 구글 시트 연결에 실패했습니다. (보안키 확인 필요)")

# --- [4. 실시간 점검 관리 대시보드] ---
def show_realtime_dashboard():
    st.divider()
    st.header("📊 실시간 점검 관리 대시보드")
    
    sheet = connect_google_sheet()
    if sheet:
        try:
            data = sheet.get_all_records()
            if not data:
                st.info("현재 구글 시트에 분석할 데이터가 없습니다.")
                return
            
            df = pd.DataFrame(data)
            df['일자'] = pd.to_datetime(df['일자'])
            df['월'] = df['일자'].dt.to_period('M').astype(str)
            
            # 분석 월 선택
            selected_month = st.selectbox("📅 분석 월 선택", sorted(df['월'].unique(), reverse=True), key="dashboard_month")
            month_df = df[df['월'] == selected_month]

            # (1) 건물별 점검 완료율 현황
            st.subheader(f"🏁 {selected_month} 건물별 점검 완료율")
            b_list = list(BUILDING_CONFIG.keys())
            for i in range(0, len(b_list), 3):
                m_cols = st.columns(3)
                for j, bname in enumerate(b_list[i:i+3]):
                    targets = BUILDING_CONFIG[bname]
                    # 구역 텍스트 매칭 및 층 정보 추출
                    b_records = month_df[month_df['구역'].str.contains(bname, na=False)]
                    done_floors = b_records['구역'].apply(lambda x: x.split()[-1] if ' ' in x else "").unique()
                    
                    done_count = len([f for f in done_floors if f in targets])
                    total_count = len(targets)
                    completion_rate = (done_count / total_count) * 100
                    
                    with m_cols[j]:
                        st.metric(label=f"📍 {bname}", value=f"{completion_rate:.1f}%", delta=f"{done_count}/{total_count} 층")
                        missing = sorted(list(set(targets) - set(done_floors)))
                        if missing:
                            st.error(f"🚩 미점검: {', '.join(missing)}")
                        else:
                            st.success("✅ 모든 층 점검 완료")

            # (2) 건물별 양호/불량 통계 그래프
            st.subheader("📈 건물별 설비 상태 합계")
            avail_cols = [c for c in TOTAL_ITEMS if c in month_df.columns]
            melted = month_df.melt(id_vars=['구역'], value_vars=avail_cols, var_name='설비명', value_name='상태')
            melted['건물'] = melted['구역'].apply(lambda x: x.split(' ')[0] if ' ' in x else x)

            fig = px.histogram(melted, x="건물", color="상태", barmode="group",
                               color_discrete_map={'양호': '#2ECC71', '불량': '#E74C3C'},
                               category_orders={"상태": ["양호", "불량"]},
                               title=f"{selected_month} 건물별 설비 상태 분포")
            
            fig.update_layout(yaxis_title="점검 항목 수", xaxis_title="건물명", legend_title="상태")
            st.plotly_chart(fig, use_container_width=True)

        except Exception as e:
            st.error(f"데이터 분석 중 오류 발생: {e}")

# 페이지 최하단에 대시보드 출력
show_realtime_dashboard()